"""All of Juile's capabilities: shell, files, desktop control, web search,
deep research, MCP (Composio + Zapier), and skills.

Every tool is ``async def`` and returns a plain string that is fed back into the
agent loop. Blocking work (shell, pyautogui) runs in a thread to stay friendly to
the event loop on Windows.
"""
import asyncio
import difflib
import json
import os
import re
import subprocess
import webbrowser
from datetime import datetime
from pathlib import Path

import httpx

from . import config, overlay

MAX_RESULT = 7000  # chars of tool output fed back to the model


def _clip(text: str, limit: int = MAX_RESULT) -> str:
    text = text or ""
    if len(text) > limit:
        return text[:limit] + f"\n…[truncated {len(text) - limit} chars]"
    return text


def _resolve(path: str) -> Path:
    p = Path(path)
    if not p.is_absolute():
        p = config.WORKSPACE / p
    return p


# Recent file diffs, so the UI can show the changed lines on click.
_DIFFS: dict[str, str] = {}


def _record_diff(display: str, old: str, new: str):
    diff = list(difflib.unified_diff(old.splitlines(), new.splitlines(), lineterm="", n=2))
    added = sum(1 for l in diff if l.startswith("+") and not l.startswith("+++"))
    removed = sum(1 for l in diff if l.startswith("-") and not l.startswith("---"))
    _DIFFS[display] = "\n".join(diff[:500])
    if len(_DIFFS) > 60:
        for k in list(_DIFFS)[:20]:
            _DIFFS.pop(k, None)
    return added, removed


# --------------------------------------------------------------------------- #
# Local machine tools
# --------------------------------------------------------------------------- #
def _shell_sync(command: str, timeout: int) -> str:
    res = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", command],
        capture_output=True, text=True, timeout=timeout, cwd=str(config.WORKSPACE),
    )
    out = (res.stdout or "") + (res.stderr or "")
    return out.strip() or f"(no output, exit code {res.returncode})"


async def shell(args: dict) -> str:
    cmd = args.get("command", "")
    if not cmd:
        return "Error: no command provided."
    try:
        return _clip(await asyncio.to_thread(_shell_sync, cmd, int(args.get("timeout", 120))))
    except subprocess.TimeoutExpired:
        return "Error: command timed out."
    except Exception as e:
        return f"Error running command: {e}"


async def python_exec(args: dict) -> str:
    code = args.get("code", "")
    if not code:
        return "Error: no code provided."
    f = config.WORKSPACE / f"_snippet_{datetime.now():%H%M%S%f}.py"
    f.write_text(code, encoding="utf-8")
    try:
        res = await asyncio.to_thread(
            lambda: subprocess.run(["python", str(f)], capture_output=True, text=True,
                                   timeout=int(args.get("timeout", 120)), cwd=str(config.WORKSPACE))
        )
        return _clip(((res.stdout or "") + (res.stderr or "")).strip() or "(no output)")
    except Exception as e:
        return f"Error: {e}"
    finally:
        try:
            f.unlink()
        except Exception:
            pass


async def write_file(args: dict) -> str:
    path = args.get("path")
    if not path:
        return "Error: 'path' required."
    p = _resolve(path)
    existed = p.exists()
    old = p.read_text(encoding="utf-8", errors="ignore") if existed else ""
    new = args.get("content", "")
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(new, encoding="utf-8")
    added, removed = _record_diff(path, old, new)
    verb = "Updated" if existed else "Created"
    return f"{verb} {p.name} (+{added} -{removed}). [[DIFF:{path}:{added}:{removed}]]"


async def read_file(args: dict) -> str:
    p = _resolve(args.get("path", ""))
    if not p.exists():
        return f"Error: {p} not found."
    try:
        return _clip(p.read_text(encoding="utf-8", errors="ignore"))
    except Exception as e:
        return f"Error reading file: {e}"


async def list_dir(args: dict) -> str:
    p = _resolve(args.get("path", "."))
    if not p.exists():
        return f"Error: {p} not found."
    items = []
    for child in sorted(p.iterdir()):
        items.append(("DIR  " if child.is_dir() else "FILE ") + child.name)
    return _clip(f"{p}\n" + "\n".join(items) or "(empty)")


async def open_app(args: dict) -> str:
    target = args.get("target", "")
    if not target:
        return "Error: 'target' required (app name, file path, or URL)."
    if target.startswith(("http://", "https://")):
        webbrowser.open(target)
        return f"Opened URL: {target}"
    try:
        await asyncio.to_thread(
            lambda: subprocess.Popen(["cmd", "/c", "start", "", target], shell=False)
        )
        return f"Launched: {target}"
    except Exception as e:
        return f"Error launching {target}: {e}"


def _coord(v):
    """Coerce a coordinate to int; return None if not provided/parseable."""
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except Exception:
        return None


def _computer_sync(args: dict) -> str:
    import pyautogui
    pyautogui.FAILSAFE = False
    action = args.get("action", "")
    x, y = _coord(args.get("x")), _coord(args.get("y"))
    w, h = pyautogui.size()
    if action in ("move", "click", "double_click", "right_click", "drag") and x is not None:
        x = max(0, min(w - 1, x))
        y = max(0, min(h - 1, y if y is not None else 0))
    # keep Juile's blue cursor in sync with where it is about to act
    if action in ("move", "click", "double_click", "right_click", "drag") and x is not None and y is not None:
        try:
            overlay.set_pos(x, y)
        except Exception:
            pass
    if action == "move":
        if x is None or y is None:
            return "Error: move needs integer x and y."
        pyautogui.moveTo(x, y, duration=0.25)
        return f"Moved cursor to ({x},{y}). Screen is {w}x{h}."
    if action == "click":
        if x is not None and y is not None:
            pyautogui.moveTo(x, y, duration=0.2)
        try:
            overlay.pulse()
        except Exception:
            pass
        pyautogui.click()
        return f"Clicked at ({pyautogui.position().x},{pyautogui.position().y})."
    if action == "double_click":
        if x is not None and y is not None:
            pyautogui.moveTo(x, y, duration=0.2)
        try:
            overlay.pulse()
        except Exception:
            pass
        pyautogui.doubleClick()
        return f"Double-clicked at ({pyautogui.position().x},{pyautogui.position().y})."
    if action == "right_click":
        if x is not None and y is not None:
            pyautogui.moveTo(x, y, duration=0.2)
        try:
            overlay.pulse()
        except Exception:
            pass
        pyautogui.rightClick()
        return f"Right-clicked at ({pyautogui.position().x},{pyautogui.position().y})."
    if action == "drag":
        if x is None or y is None:
            return "Error: drag needs destination x and y."
        pyautogui.dragTo(x, y, duration=0.45, button="left")
        return f"Dragged to ({x},{y})."
    if action == "scroll":
        pyautogui.scroll(_coord(args.get("amount")) or -300); return "Scrolled."
    if action == "type":
        pyautogui.typewrite(str(args.get("text", "")), interval=0.02); return "Typed text."
    if action == "press":
        pyautogui.press(str(args.get("text", "enter"))); return f"Pressed {args.get('text')}."
    if action == "hotkey":
        keys = [k.strip() for k in str(args.get("text", "")).replace("+", ",").split(",") if k.strip()]
        pyautogui.hotkey(*keys); return f"Pressed hotkey {'+'.join(keys)}."
    if action == "screenshot":
        path = config.WORKSPACE / f"screen_{datetime.now():%H%M%S}.png"
        pyautogui.screenshot(str(path))
        return f"Screenshot captured. Screen {w}x{h}. Cursor at {pyautogui.position()}. [[SHOT:{path.name}]]"
    if action == "position":
        return f"Cursor at {pyautogui.position()}. Screen size is {w}x{h}."
    if action in ("key_down", "key_up", "hold"):
        import time as _time
        keys = [k.strip() for k in str(args.get("text", "")).replace("+", ",").split(",") if k.strip()]
        if action == "key_down":
            for k in keys:
                pyautogui.keyDown(k)
            return f"Holding {'+'.join(keys) or '(nothing)'} down."
        if action == "key_up":
            if not keys:
                for k in ("shift", "ctrl", "alt", "win", "shiftleft", "ctrlleft", "altleft"):
                    pyautogui.keyUp(k)
                return "Released held keys."
            for k in keys:
                pyautogui.keyUp(k)
            return f"Released {'+'.join(keys)}."
        dur = _coord(args.get("amount")) or _coord(args.get("duration")) or 800
        for k in keys:
            pyautogui.keyDown(k)
        _time.sleep(min(15000, max(0, dur)) / 1000.0)
        for k in reversed(keys):
            pyautogui.keyUp(k)
        return f"Held {'+'.join(keys)} for {dur}ms."
    if action == "wait":
        import time as _time
        ms = _coord(args.get("amount")) or 500
        _time.sleep(min(10000, max(0, ms)) / 1000.0)
        return f"Waited {ms}ms."
    return f"Unknown computer action: {action}. Valid: move, click, double_click, right_click, drag, scroll, type, press, hotkey, key_down, key_up, hold, wait, screenshot, position."


async def computer(args: dict) -> str:
    try:
        overlay.show()   # ensure the blue cursor is visible for any computer action
        return await asyncio.to_thread(_computer_sync, args)
    except ImportError:
        return "Error: pyautogui not installed. Run: pip install pyautogui"
    except Exception as e:
        return f"Computer-control error: {e}"


def _window_sync(args: dict) -> str:
    try:
        import pygetwindow as gw
        import pyautogui
    except Exception:
        return "Window control needs pygetwindow. Run: pip install pygetwindow"
    action = (args.get("action") or "").lower()
    title = (args.get("title") or "").strip()
    sw, sh = pyautogui.size()
    if action == "list":
        titles = [t for t in gw.getAllTitles() if t.strip()]
        return ("Open windows:\n" + "\n".join(f"- {t}" for t in titles[:40])) if titles else "(no windows)"
    win = None
    try:
        if title:
            matches = gw.getWindowsWithTitle(title)
            win = matches[0] if matches else None
        else:
            win = gw.getActiveWindow()
    except Exception:
        win = None
    if not win:
        return f"No window found{(' titled ' + repr(title)) if title else ''}. Try action='list' first."
    try:
        if action in ("focus", "activate"):
            try:
                win.activate()
            except Exception:
                win.minimize(); win.restore()
        elif action == "minimize":
            win.minimize()
        elif action == "maximize":
            win.maximize()
        elif action == "restore":
            win.restore()
        elif action == "close":
            win.close()
        elif action == "move":
            win.moveTo(_coord(args.get("x")) or 0, _coord(args.get("y")) or 0)
        elif action == "resize":
            win.resizeTo(_coord(args.get("width")) or win.width, _coord(args.get("height")) or win.height)
        elif action in ("snap_left", "left"):
            win.restore(); win.moveTo(0, 0); win.resizeTo(sw // 2, sh)
        elif action in ("snap_right", "right"):
            win.restore(); win.moveTo(sw // 2, 0); win.resizeTo(sw // 2, sh)
        elif action in ("snap_top", "top"):
            win.restore(); win.moveTo(0, 0); win.resizeTo(sw, sh // 2)
        elif action in ("snap_bottom", "bottom"):
            win.restore(); win.moveTo(0, sh // 2); win.resizeTo(sw, sh // 2)
        else:
            return f"Unknown window action: {action}. Valid: list, focus, minimize, maximize, restore, close, move, resize, snap_left, snap_right, snap_top, snap_bottom."
        return f"Window '{getattr(win, 'title', '')[:50]}': {action} done."
    except Exception as e:
        return f"Window error: {e}"


async def window(args: dict) -> str:
    try:
        return await asyncio.to_thread(_window_sync, args)
    except Exception as e:
        return f"Window-control error: {e}"


# --------------------------------------------------------------------------- #
# Web search + deep research (Tavily)
# --------------------------------------------------------------------------- #
async def _tavily(query: str, max_results: int = 6, advanced: bool = False, raw: bool = False) -> dict:
    key = config.tavily_key()
    if not key:
        return {"error": "Tavily key not set (add it in Settings -> Connectors)."}
    payload = {
        "api_key": key,
        "query": query,
        "max_results": max_results,
        "search_depth": "advanced" if advanced else "basic",
        "include_answer": True,
        "include_raw_content": raw,
        "include_images": True,
    }
    async with httpx.AsyncClient(timeout=60) as c:
        r = await c.post("https://api.tavily.com/search", json=payload)
        if r.status_code >= 400:
            return {"error": f"Tavily HTTP {r.status_code}: {r.text[:300]}"}
        return r.json()


async def web_search(args: dict) -> str:
    data = await _tavily(args.get("query", ""), int(args.get("max_results", 6)))
    if "error" in data:
        return f"Error: {data['error']}"
    lines = []
    if data.get("answer"):
        lines.append(f"Quick answer: {data['answer']}\n")
    for i, r in enumerate(data.get("results", []), 1):
        lines.append(f"[{i}] {r.get('title')}\n    {r.get('url')}\n    {(r.get('content') or '')[:300]}")
    imgs = [im.get("url") if isinstance(im, dict) else im for im in (data.get("images") or [])]
    imgs = [u for u in imgs if u]
    if imgs:
        lines.append("\nImages you can embed inline with ![](url) to illustrate the answer:\n" + "\n".join(imgs[:6]))
    return _clip("\n".join(lines) or "No results.")


async def deep_research(args: dict) -> str:
    """Fan out many searches, dedupe sources, return a cited bundle for synthesis.

    The model supplies `subqueries` (the decomposition). We search each in
    parallel with advanced depth, dedupe by URL, and hand back a large bundle of
    sources the model then synthesises into a report with [n] citations.
    """
    topic = args.get("topic", "")
    subs = args.get("subqueries") or []
    if not subs:
        subs = [topic, f"{topic} overview", f"{topic} latest 2026",
                f"{topic} statistics data", f"{topic} criticism risks",
                f"{topic} how it works", f"{topic} comparison alternatives"]
    subs = subs[:12]
    per = int(args.get("max_per_query", 8))

    results = await asyncio.gather(*[_tavily(q, per, advanced=True) for q in subs],
                                   return_exceptions=True)
    seen, sources = set(), []
    for data in results:
        if isinstance(data, Exception) or not isinstance(data, dict) or "error" in data:
            continue
        for r in data.get("results", []):
            url = r.get("url")
            if not url or url in seen:
                continue
            seen.add(url)
            sources.append(r)

    if not sources:
        return "Deep research found no sources (check Tavily key / connectivity)."

    out = [f"DEEP RESEARCH BUNDLE — {len(sources)} unique sources across {len(subs)} queries.",
           "Synthesise a structured, cited report. Cite sources as [n].\n"]
    for i, r in enumerate(sources, 1):
        snippet = (r.get("content") or "")[:400]
        out.append(f"[{i}] {r.get('title')}\n{r.get('url')}\n{snippet}\n")
    return _clip("\n".join(out), limit=14000)


# --------------------------------------------------------------------------- #
# MCP — Composio + Zapier + Higgsfield (streamable HTTP)
# --------------------------------------------------------------------------- #
def _composio_target():
    c = config.mcp_creds("composio")
    return c["url"], ({"x-consumer-api-key": c["api_key"]} if c["api_key"] else {})


def _zapier_target():
    return config.mcp_creds("zapier")["url"], {}


def _higgsfield_target():
    c = config.mcp_creds("higgsfield")
    return c["url"], ({"Authorization": f"Bearer {c['api_key']}"} if c["api_key"] else {})


_MCP_TARGETS = {"composio": _composio_target, "zapier": _zapier_target, "higgsfield": _higgsfield_target}
_mcp_tool_cache: dict[str, list] = {}
_HF_CALLBACK_PORT = 8766


# --- Higgsfield OAuth (browser flow on first use; tokens cached on disk) ------ #
class _FileTokenStorage:
    """Persists OAuth tokens + dynamic client registration to a JSON file."""

    def __init__(self, path: Path):
        self.path = path

    def _read(self) -> dict:
        try:
            return json.loads(self.path.read_text(encoding="utf-8"))
        except Exception:
            return {}

    def _write(self, data: dict):
        self.path.write_text(json.dumps(data), encoding="utf-8")

    async def get_tokens(self):
        from mcp.shared.auth import OAuthToken
        d = self._read().get("tokens")
        return OAuthToken(**d) if d else None

    async def set_tokens(self, tokens):
        d = self._read(); d["tokens"] = tokens.model_dump(exclude_none=True); self._write(d)

    async def get_client_info(self):
        from mcp.shared.auth import OAuthClientInformationFull
        d = self._read().get("client")
        return OAuthClientInformationFull(**d) if d else None

    async def set_client_info(self, info):
        d = self._read(); d["client"] = info.model_dump(exclude_none=True, mode="json"); self._write(d)


async def _hf_redirect(url: str):
    webbrowser.open(url)


async def _hf_callback():
    """One-shot local server that catches the OAuth redirect and returns (code, state)."""
    import http.server
    import threading
    import urllib.parse

    result, done = {}, threading.Event()

    class H(http.server.BaseHTTPRequestHandler):
        def do_GET(self):
            result.update(dict(urllib.parse.parse_qsl(urllib.parse.urlparse(self.path).query)))
            self.send_response(200); self.send_header("Content-Type", "text/html"); self.end_headers()
            self.wfile.write(b"<body style='background:#000;color:#fff;font-family:sans-serif;"
                             b"display:grid;place-items:center;height:100vh;margin:0'>"
                             b"<h2>Higgsfield connected. Close this tab and return to Juile.</h2></body>")
            done.set()

        def log_message(self, *a):
            pass

    srv = http.server.HTTPServer(("127.0.0.1", _HF_CALLBACK_PORT), H)
    threading.Thread(target=srv.handle_request, daemon=True).start()
    await asyncio.to_thread(done.wait, 300)
    srv.server_close()
    return result.get("code", ""), result.get("state")


def _higgsfield_auth():
    """Build an OAuth provider for Higgsfield, or None if a static key is set / unavailable."""
    creds = config.mcp_creds("higgsfield")
    if creds["api_key"]:
        return None
    try:
        from mcp.client.auth import OAuthClientProvider
        from mcp.shared.auth import OAuthClientMetadata
        return OAuthClientProvider(
            server_url=creds["url"],
            client_metadata=OAuthClientMetadata(
                client_name="Juile",
                redirect_uris=[f"http://localhost:{_HF_CALLBACK_PORT}/callback"],
                grant_types=["authorization_code", "refresh_token"],
                response_types=["code"],
                token_endpoint_auth_method="none",
            ),
            storage=_FileTokenStorage(config.WORKSPACE / "higgsfield_oauth.json"),
            redirect_handler=_hf_redirect,
            callback_handler=_hf_callback,
        )
    except Exception:
        return None


def _mcp_endpoint(server: str):
    """Return (url, headers, auth) for a server, or (None, msg, None) on error."""
    if server in _MCP_TARGETS:
        url, headers = _MCP_TARGETS[server]()
        auth = _higgsfield_auth() if server == "higgsfield" else None
        return url, headers, auth
    from . import store
    conn = next((c for c in store.load().get("connectors", []) if c.get("name", "").lower() == server.lower()), None)
    if not conn:
        names = ", ".join(list(_MCP_TARGETS) + [c.get("name", "") for c in store.load().get("connectors", [])])
        return None, f"Error: unknown MCP server '{server}'. Available: {names}", None
    headers = {conn["header_name"]: conn["header_value"]} if conn.get("header_name") else {}
    return conn.get("url", ""), headers, None


HF_IMAGE_MODEL = "nano_banana_pro"   # the ONLY image model Juile is allowed to use
_URL_RE = re.compile(r'https?://[^\s"\'<>)\]}]+', re.I)
_UUID_RE = re.compile(r'[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}')


def _content_text(result) -> str:
    return "\n".join(getattr(b, "text", None) or str(b) for b in (getattr(result, "content", None) or []))


def _first_image_url(text: str):
    urls = _URL_RE.findall(text or "")
    for u in urls:
        if u.lower().split("?")[0].endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")):
            return u
    for u in urls:                       # CDN URLs without a file extension
        if any(h in u.lower() for h in ("higgsfield", "cloudfront", "amazonaws", "storage", "cdn", "blob")):
            return u
    return None


def _url_for_job(text: str, jid: str):
    """Find the image URL that belongs to a specific job id (avoids picking a stale one)."""
    i = (text or "").find(jid or "\0")
    return _first_image_url(text[i:i + 4000]) if i != -1 else None


async def _higgsfield_image(session, tool: str, arguments: dict) -> str:
    """Generate an image with nano_banana_pro and WAIT here until the URL is ready,
    so the whole thing costs the agent exactly one step (no polling spam)."""
    arguments = dict(arguments or {})
    if tool == "generate_image":
        params = dict(arguments.get("params") or {})
        params["model"] = HF_IMAGE_MODEL          # force Nano Banana Pro, always
        arguments["params"] = params
    result = await session.call_tool(tool, arguments)
    text = _content_text(result)
    url = _first_image_url(text)
    m = _UUID_RE.search(text)
    jid = m.group(0) if m else None
    if not url and jid:                            # async job — poll THIS job to completion here
        for _ in range(24):                        # up to ~96s
            await asyncio.sleep(4)
            try:
                probe = await session.call_tool("show_generations", {"type": "image", "size": 10})
            except Exception:
                break
            t2 = _content_text(probe)
            url = _url_for_job(t2, jid)
            if url:
                break
    if url:
        text += f"\n[[IMG:{url}]]"
    return _clip(text or "(no content returned)")


async def _mcp_run(server: str, op: str, tool: str = "", arguments: dict | None = None):
    try:
        from mcp import ClientSession
        from mcp.client.streamable_http import streamablehttp_client
    except ImportError:
        return "Error: mcp package not installed. Run: pip install mcp"
    url, headers, auth = _mcp_endpoint(server)
    if url is None:
        return headers  # error message
    if not url:
        return f"Error: {server} is not connected. Add its URL/key in Settings -> Connectors."
    if server == "composio" and not config.mcp_creds("composio")["api_key"]:
        return ("Composio isn't connected yet. Open Settings -> Connectors, paste your Composio API key "
                "(and your personal MCP URL if you have one), Save, then ask me again.")
    if server == "zapier" and not config.mcp_creds("zapier")["url"]:
        return "Zapier isn't connected. Paste your Zapier MCP URL in Settings -> Connectors, then try again."
    try:
        async with streamablehttp_client(url, headers=headers, auth=auth) as (read, write, _):
            async with ClientSession(read, write) as session:
                await session.initialize()
                if op == "list":
                    tools = (await session.list_tools()).tools
                    _mcp_tool_cache[server] = [t.name for t in tools]
                    return "\n".join(f"- {t.name}: {(t.description or '')[:160]}" for t in tools) or "(no tools)"
                if server == "higgsfield" and tool in ("generate_image", "edit_image"):
                    return await _higgsfield_image(session, tool, arguments or {})
                result = await session.call_tool(tool, arguments or {})
                parts = [getattr(b, "text", None) or str(b) for b in (result.content or [])]
                return _clip("\n".join(parts) or "(no content returned)")
    except Exception as e:
        msg = str(e)
        if server == "higgsfield" and ("auth" in msg.lower() or "401" in msg or "oauth" in msg.lower()):
            return ("Higgsfield needs you to sign in. A browser tab should have opened — approve access, "
                    "then run this again. (If no tab opened, set the Higgsfield key in Settings -> Connectors.)")
        if server == "composio" and any(k in msg.lower() for k in ("401", "403", "404", "auth", "unauthor", "not found", "invalid")):
            return ("Couldn't reach Composio — double-check your Composio API key and MCP URL in "
                    f"Settings -> Connectors. ({msg[:160]})")
        return f"MCP ({server}) error: {msg}"


async def mcp(args: dict) -> str:
    server = (args.get("server") or "").lower()
    op = (args.get("op") or "call").lower()
    return await _mcp_run(server, op, args.get("tool", ""), args.get("arguments") or {})


async def connector_auth(name: str) -> dict:
    """Authenticate / verify a connector right from Settings. Runs the real MCP
    connect handshake — which kicks off Higgsfield's browser OAuth, validates a
    Composio key, and proves a custom server reachable — and reports the result.
    Returns {ok, message} for the Auth button to show inline."""
    name = (name or "").strip()
    if not name:
        return {"ok": False, "message": "No connector specified."}
    low = name.lower()
    # Tavily is a plain search key (not an MCP server) — validate it directly.
    if low == "tavily":
        if not config.tavily_key():
            return {"ok": False, "message": "No Tavily key saved yet — paste it above and Save first."}
        data = await _tavily("ping", max_results=1)
        if isinstance(data, dict) and "error" in data:
            return {"ok": False, "message": f"Tavily rejected that key: {str(data['error'])[:160]}"}
        return {"ok": True, "message": "Connected — web search & deep research are live."}
    try:
        text = (await _mcp_run(low, "list")) or ""
    except Exception as e:
        return {"ok": False, "message": f"Couldn't connect: {str(e)[:200]}"}
    low_t = text.lower()
    if text.lstrip().startswith("- ") or "\n- " in text:
        n = text.count("- ")
        return {"ok": True, "message": f"Connected — {n} tool{'' if n == 1 else 's'} available."}
    if any(k in low_t for k in ("sign in", "browser", "approve access", "a browser tab")):
        return {"ok": True, "message": text.strip()[:400]}        # OAuth flow opened in the browser
    if any(k in low_t for k in ("error", "not connected", "isn't connected", "unknown mcp", "couldn't reach")):
        return {"ok": False, "message": text.strip()[:400]}
    return {"ok": True, "message": text.strip()[:400] or "Connected."}


# --------------------------------------------------------------------------- #
# GitHub — read & interact with the user's repos via the gh CLI
# --------------------------------------------------------------------------- #
async def github(args: dict) -> str:
    """Run a GitHub CLI (`gh`) command to read or act on the user's repos:
    view files, list/read issues & PRs, open PRs, clone, etc. Defaults to the
    active repo (Settings/composer) when the command needs one."""
    import shlex
    import shutil
    gh = shutil.which("gh") or shutil.which("gh.exe")
    if not gh:
        return "Error: GitHub CLI (gh) not installed. Ask the user to install it from cli.github.com and sign in."
    cmd = args.get("command") or args.get("args") or ""
    if isinstance(cmd, list):
        parts = [str(x) for x in cmd]
    else:
        cmd = str(cmd).strip()
        if cmd.lower().startswith("gh "):
            cmd = cmd[3:]
        try:
            parts = shlex.split(cmd, posix=False)
        except Exception:
            parts = cmd.split()
    if not parts:
        return "Error: no gh command given."
    try:
        from . import store
        repo = (store.load().get("github_repo") or "").strip()
    except Exception:
        repo = ""
    argv = [gh] + parts
    try:
        res = await asyncio.to_thread(
            lambda: subprocess.run(argv, capture_output=True, text=True,
                                   timeout=int(args.get("timeout", 120)), cwd=str(config.WORKSPACE)))
        out = ((res.stdout or "") + (res.stderr or "")).strip()
        hint = f"\n(active repo: {repo})" if repo and "-R" not in parts and "--repo" not in parts else ""
        return _clip((out or f"(no output, exit {res.returncode})") + hint)
    except subprocess.TimeoutExpired:
        return "Error: gh command timed out."
    except Exception as e:
        return f"Error running gh: {e}"


# --------------------------------------------------------------------------- #
# Skills
# --------------------------------------------------------------------------- #
async def skill(args: dict) -> str:
    op = (args.get("op") or "list").lower()
    if op == "list":
        out = []
        for f in sorted(config.SKILLS_DIR.glob("*.md")):
            first = f.read_text(encoding="utf-8", errors="ignore").splitlines()
            desc = next((l.strip("# ").strip() for l in first if l.strip()), "")
            out.append(f"- {f.stem}: {desc[:140]}")
        return "\n".join(out) or "(no skills installed)"
    name = args.get("name", "")
    f = config.SKILLS_DIR / f"{name}.md"
    if not f.exists():
        return f"Error: skill '{name}' not found."
    return _clip(f.read_text(encoding="utf-8", errors="ignore"))


# --------------------------------------------------------------------------- #
# Files, spreadsheets, memory
# --------------------------------------------------------------------------- #
async def create_file(args: dict) -> str:
    path = args.get("path") or args.get("filename")
    if not path:
        return "Error: 'path' required."
    p = _resolve(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(args.get("content", ""), encoding="utf-8")
    return f"Created {p.name}. [[FILE:{p.name}]]"


def _xlsx_sync(args: dict) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    name = args.get("filename", "spreadsheet.xlsx")
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    p = _resolve(name)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (args.get("sheet") or "Sheet1")[:31]
    cols = args.get("columns", [])
    rows = args.get("rows", [])
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F2937")
    band = PatternFill("solid", fgColor="F3F4F6")
    if cols:
        ws.append([str(c) for c in cols])
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = head_fill
            c.alignment = Alignment(vertical="center")
            c.border = border
        ws.row_dimensions[1].height = 22
    for i, r in enumerate(rows):
        ws.append(list(r))
        for c in ws[ws.max_row]:
            c.border = border
            if i % 2 == 1:
                c.fill = band
    # auto width
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(48, width + 4)
    if cols:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    wb.save(p)
    return f"Spreadsheet '{p.name}' created with {len(rows)} rows. [[FILE:{p.name}]]"


async def create_xlsx(args: dict) -> str:
    try:
        return await asyncio.to_thread(_xlsx_sync, args)
    except ImportError:
        return "Error: openpyxl not installed."
    except Exception as e:
        return f"Spreadsheet error: {e}"


async def remember(args: dict) -> str:
    from . import store
    text = args.get("text", "").strip()
    if not text:
        return "Error: nothing to remember."
    store.add_memory(text)
    return f"Saved to memory: {text}"


async def make_dir(args: dict) -> str:
    path = args.get("path", "")
    if not path:
        return "Error: 'path' required."
    p = _resolve(path)
    try:
        p.mkdir(parents=True, exist_ok=True)
        return f"Created directory {p}"
    except Exception as e:
        return f"Error creating directory: {e}"


async def move_path(args: dict) -> str:
    import shutil
    src = _resolve(args.get("src") or args.get("from") or "")
    dst = _resolve(args.get("dst") or args.get("to") or "")
    if not src.exists():
        return f"Error: source {src} not found."
    try:
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(src), str(dst))
        return f"Moved {src.name} -> {dst}"
    except Exception as e:
        return f"Error moving: {e}"


async def copy_path(args: dict) -> str:
    import shutil
    src = _resolve(args.get("src") or args.get("from") or "")
    dst = _resolve(args.get("dst") or args.get("to") or "")
    if not src.exists():
        return f"Error: source {src} not found."
    try:
        dst.parent.mkdir(parents=True, exist_ok=True)
        if src.is_dir():
            shutil.copytree(str(src), str(dst), dirs_exist_ok=True)
        else:
            shutil.copy2(str(src), str(dst))
        return f"Copied {src.name} -> {dst}"
    except Exception as e:
        return f"Error copying: {e}"


async def delete_path(args: dict) -> str:
    import shutil
    if not args.get("path"):
        return "Error: 'path' required."
    p = _resolve(args.get("path"))
    if not p.exists():
        return f"{p} does not exist."
    try:
        if p.is_dir():
            shutil.rmtree(str(p))
        else:
            p.unlink()
        return f"Deleted {p}"
    except Exception as e:
        return f"Error deleting: {e}"


# --------------------------------------------------------------------------- #
# Registry
# --------------------------------------------------------------------------- #
REGISTRY = {
    "shell": shell,
    "python": python_exec,
    "write_file": write_file,
    "read_file": read_file,
    "list_dir": list_dir,
    "open_app": open_app,
    "computer": computer,
    "web_search": web_search,
    "deep_research": deep_research,
    "mcp": mcp,
    "skill": skill,
    "create_file": create_file,
    "create_xlsx": create_xlsx,
    "remember": remember,
    "make_dir": make_dir,
    "move_path": move_path,
    "copy_path": copy_path,
    "delete_path": delete_path,
    "window": window,
    "github": github,
}

# Human-readable specs injected into the system prompt.
TOOL_SPECS = [
    ("shell", "Run a PowerShell command on the user's Windows PC.", '{"command": "Get-Date"}'),
    ("python", "Execute a Python snippet and capture its output.", '{"code": "print(2+2)"}'),
    ("write_file", "Write/overwrite a text file (code, docs, anything).", '{"path": "app/main.py", "content": "..."}'),
    ("read_file", "Read a text file.", '{"path": "notes.txt"}'),
    ("list_dir", "List a directory.", '{"path": "."}'),
    ("open_app", "Open an app, file, or URL (e.g. notepad, calc, a path, https://...).", '{"target": "notepad"}'),
    ("computer", "Drive Juile's own blue cursor: mouse/keyboard/screen. Screenshot first, then integer x,y. action: move|click|double_click|right_click|drag|scroll|type|press|hotkey|key_down|key_up|hold|wait|screenshot|position. HOLD a keybind with action=hold (text='w', amount=ms); key_down/key_up to hold across steps.", '{"action": "click", "x": 500, "y": 300}'),
    ("window", "Organize windows/pages: list|focus|minimize|maximize|restore|close|move|resize|snap_left|snap_right|snap_top|snap_bottom. Target by title substring (or the active window).", '{"action": "snap_left", "title": "Chrome"}'),
    ("web_search", "Quick web search (Tavily).", '{"query": "weather in Tokyo", "max_results": 6}'),
    ("deep_research", "Massive multi-query research. Provide your own subqueries for breadth; returns a big cited source bundle to synthesise.", '{"topic": "solid state batteries", "subqueries": ["...","..."], "max_per_query": 8}'),
    ("mcp", "Operate real apps. server=composio (Gmail/Calendar/Notion/Slack/500+), zapier, or higgsfield (generate/edit images, video, and slide decks). Composio flow: ONE COMPOSIO_SEARCH_TOOLS, then COMPOSIO_MULTI_EXECUTE_TOOL. Higgsfield: op=list once to see tools, then call the generate tool.", '{"server": "higgsfield", "op": "call", "tool": "generate_image", "arguments": {"prompt": "a happy dog"}}'),
    ("github", "Read & act on the user's GitHub repos via the gh CLI (view files, issues, PRs, clone, open PRs). Defaults to the active repo.", '{"command": "repo view owner/name"}'),
    ("skill", "List or read a local skill. op=list|read.", '{"op": "list"}'),
    ("create_file", "Create a file with text content; it shows as a file card with an icon.", '{"path": "notes.md", "content": "..."}'),
    ("create_xlsx", "Create a BEAUTIFUL styled Excel spreadsheet (headers, banded rows, filters).", '{"filename": "budget.xlsx", "columns": ["Item","Cost"], "rows": [["Rent",1200]]}'),
    ("remember", "Save a durable fact about the user to long-term memory.", '{"text": "User prefers dark mode and concise answers."}'),
    ("make_dir", "Create a folder (and any parent folders).", '{"path": "src/components"}'),
    ("move_path", "Move or rename a file or folder.", '{"src": "a.txt", "dst": "b.txt"}'),
    ("copy_path", "Copy a file or folder.", '{"src": "a", "dst": "backup/a"}'),
    ("delete_path", "Delete a file or folder (recursive).", '{"path": "old.tmp"}'),
    ("make_plan", "Show a checklist plan for a heavy task (renders in chat and the Plan widget).", '{"title": "Build the site", "steps": ["Design","Code","Deploy"]}'),
    ("make_slides", "Build a slide deck in ONE call (shows 'Generating Slide... n/n'). Each slide: title + bullets; vivid gradient background is automatic. Add image_prompt only if the user wants real photos.", '{"title": "My Deck", "slides": [{"title": "Intro", "bullets": ["point a","point b"]}]}'),
    ("ask_user", "Ask the user structured questions with multiple-choice answers before acting. Use when you need a decision.", '{"title": "Pick a stack", "questions": [{"title": "Framework?", "description": "Which one?", "options": [{"title": "React", "description": "Popular"}, {"title": "Vue", "description": "Simple"}]}]}'),
    ("spawn_agents", "Summon parallel sub-agents that work at once and share notes (e.g. one searches, one codes, one checks email).", '{"agents": [{"role": "researcher", "task": "find competitors"}, {"role": "writer", "task": "draft the intro"}]}'),
]


def _fn(name, desc, props, required=()):
    return {"type": "function", "function": {
        "name": name, "description": desc,
        "parameters": {"type": "object", "properties": props, "required": list(required)},
    }}


# OpenAI-style function schemas for native tool calling.
TOOLS_SCHEMA = [
    _fn("shell", "Run a PowerShell command on the user's Windows PC and return its output.",
        {"command": {"type": "string", "description": "PowerShell command"}}, ["command"]),
    _fn("python", "Execute a Python snippet and capture stdout/stderr.",
        {"code": {"type": "string"}}, ["code"]),
    _fn("write_file", "Create or overwrite a text file (code, docs, anything).",
        {"path": {"type": "string"}, "content": {"type": "string"}}, ["path", "content"]),
    _fn("read_file", "Read a text file.", {"path": {"type": "string"}}, ["path"]),
    _fn("list_dir", "List a directory.", {"path": {"type": "string"}}, ["path"]),
    _fn("open_app", "Open an app, file, or URL (e.g. chrome, notepad, calc, a path, https://...).",
        {"target": {"type": "string"}}, ["target"]),
    _fn("computer", "Drive Juile's own blue cursor over the real desktop: mouse, keyboard, and screen. Screenshot first, then integer x,y. key_down/key_up/hold let you HOLD keybinds (text='shift' or 'w'; hold uses amount=ms). wait pauses (amount=ms).",
        {"action": {"type": "string", "enum": ["move", "click", "double_click", "right_click", "drag", "scroll", "type", "press", "hotkey", "key_down", "key_up", "hold", "wait", "screenshot", "position"]},
         "x": {"type": "integer"}, "y": {"type": "integer"}, "text": {"type": "string"}, "amount": {"type": "integer"}}, ["action"]),
    _fn("window", "Organize the desktop: list windows, focus/minimize/maximize/restore/close, move/resize, or snap a window left/right/top/bottom. Target by title substring (or the active window).",
        {"action": {"type": "string", "enum": ["list", "focus", "minimize", "maximize", "restore", "close", "move", "resize", "snap_left", "snap_right", "snap_top", "snap_bottom"]},
         "title": {"type": "string"}, "x": {"type": "integer"}, "y": {"type": "integer"}, "width": {"type": "integer"}, "height": {"type": "integer"}}, ["action"]),
    _fn("web_search", "Quick web search via Tavily.",
        {"query": {"type": "string"}, "max_results": {"type": "integer"}}, ["query"]),
    _fn("deep_research", "Run massive multi-query research. Provide subqueries for breadth; returns a cited source bundle.",
        {"topic": {"type": "string"}, "subqueries": {"type": "array", "items": {"type": "string"}}, "max_per_query": {"type": "integer"}}, ["topic"]),
    _fn("mcp", "Operate the user's real apps. server=composio for Gmail/Calendar/Notion/Slack/500+ apps (call tool=COMPOSIO_SEARCH_TOOLS once with arguments={query:'...'}, then tool=COMPOSIO_MULTI_EXECUTE_TOOL to run the discovered tool); server=zapier similarly; server=higgsfield to GENERATE or EDIT images, video, and slide decks (op=list once to see its tools, then call the right generate/edit tool with arguments={prompt:'...'}).",
        {"server": {"type": "string", "enum": ["composio", "zapier", "higgsfield"]}, "op": {"type": "string", "enum": ["list", "call"]},
         "tool": {"type": "string", "description": "The exact tool name to call, e.g. COMPOSIO_SEARCH_TOOLS or generate_image"},
         "arguments": {"type": "object", "additionalProperties": True}}, ["server"]),
    _fn("github", "Read & act on the user's GitHub repos via the gh CLI. Pass a gh subcommand, e.g. 'repo view owner/name', 'api repos/owner/name/contents/path', 'issue list -R owner/name', 'pr create ...', 'repo clone owner/name'. Defaults to the active repo when set.",
        {"command": {"type": "string", "description": "gh subcommand, e.g. \"repo view owner/name\" or \"api user\""}}, ["command"]),
    _fn("skill", "List or read a local skill playbook.",
        {"op": {"type": "string", "enum": ["list", "read"]}, "name": {"type": "string"}}, []),
    _fn("create_file", "Create a text/code/markdown file; it renders as a downloadable file card.",
        {"path": {"type": "string"}, "content": {"type": "string"}}, ["path", "content"]),
    _fn("create_xlsx", "Create a beautifully styled Excel spreadsheet (headers, banded rows, filters).",
        {"filename": {"type": "string"}, "columns": {"type": "array", "items": {"type": "string"}},
         "rows": {"type": "array", "items": {"type": "array"}}}, ["filename"]),
    _fn("remember", "Save a durable fact about the user to long-term memory.",
        {"text": {"type": "string"}}, ["text"]),
    _fn("make_dir", "Create a directory (and any parent directories).",
        {"path": {"type": "string"}}, ["path"]),
    _fn("move_path", "Move or rename a file or folder.",
        {"src": {"type": "string"}, "dst": {"type": "string"}}, ["src", "dst"]),
    _fn("copy_path", "Copy a file or folder (recursive for folders).",
        {"src": {"type": "string"}, "dst": {"type": "string"}}, ["src", "dst"]),
    _fn("delete_path", "Delete a file or folder (recursive). Respects Ask-permission mode.",
        {"path": {"type": "string"}}, ["path"]),
    _fn("make_plan", "Show a checklist plan for a heavy multi-step task (renders in chat and the Plan widget).",
        {"title": {"type": "string"}, "steps": {"type": "array", "items": {"type": "string"}}}, ["steps"]),
    _fn("make_slides", "Build a slide deck in ONE call (shows live 'Generating Slide... n/n'). Each slide gets a vivid gradient background automatically; add image_prompt only for real photos.",
        {"title": {"type": "string"},
         "slides": {"type": "array", "items": {"type": "object", "properties": {
             "title": {"type": "string"}, "bullets": {"type": "array", "items": {"type": "string"}},
             "body": {"type": "string"}, "image_prompt": {"type": "string"}}}}}, ["slides"]),
    _fn("ask_user", "Ask the user multiple-choice questions before acting. Use when you need a decision.",
        {"title": {"type": "string"},
         "questions": {"type": "array", "items": {"type": "object", "properties": {
             "title": {"type": "string"}, "description": {"type": "string"},
             "options": {"type": "array", "items": {"type": "object", "properties": {
                 "title": {"type": "string"}, "description": {"type": "string"}}}}}}}}, ["questions"]),
    _fn("spawn_agents", "Summon parallel sub-agents that work at once and share notes.",
        {"agents": {"type": "array", "items": {"type": "object", "properties": {
            "role": {"type": "string"}, "task": {"type": "string"}}}}}, ["agents"]),
]
